import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PySide6.QtWidgets import QFileDialog, QMessageBox


class TableExporter:

  @staticmethod
  def export_to_excel(parent_widget, table_widget, title="Selling_Report"):
    """Экспорт данных из QTableWidget в красиво оформленный Excel-файл."""
    path, _ = QFileDialog.getSaveFileName(
        parent_widget, "Сохранить в Excel", f"{title}.xlsx", "Excel Files (*.xlsx)"
    )

    if not path:
      return

    try:
      wb = Workbook()
      ws = wb.active
      ws.title = "Отчёт"

      # Стиль для заголовка
      header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
      header_fill = PatternFill(
          start_color="1E1E2E", end_color="1E1E2E", fill_type="solid"
      )
      thin_border = Border(
          left=Side(style="thin", color="CCCCCC"),
          right=Side(style="thin", color="CCCCCC"),
          top=Side(style="thin", color="CCCCCC"),
          bottom=Side(style="thin", color="CCCCCC"),
      )

      # 1. Заголовки столбцов
      headers = []
      for col in range(table_widget.columnCount()):
        header_text = table_widget.horizontalHeaderItem(col)
        headers.append(header_text.text() if header_text else f"Col {col+1}")

      ws.append(headers)

      # Оформляем шапку
      for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

      # 2. Данные строк
      for row in range(table_widget.rowCount()):
        row_data = []
        for col in range(table_widget.columnCount()):
          item = table_widget.item(row, col)

          # Если в ячейке находится ComboBox (например, Статус)
          widget = table_widget.cellWidget(row, col)
          if widget and hasattr(widget, "currentText"):
            row_data.append(widget.currentText())
          elif item is not None:
            row_data.append(item.text())
          else:
            row_data.append("")

        ws.append(row_data)

        # Применяем границы и выравнивание к строке
        current_row = ws.max_row
        for col_num in range(1, len(row_data) + 1):
          cell = ws.cell(row=current_row, column=col_num)
          cell.border = thin_border
          cell.alignment = Alignment(vertical="center")

      # Автоподбор ширины колонок
      for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

      wb.save(path)
      QMessageBox.information(
          parent_widget,
          "Успех",
          f"Файл успешно сохранён:\n{path}",
      )

    except Exception as e:
      QMessageBox.critical(
          parent_widget, "Ошибка экспорта", f"Не удалось сохранить файл:\n{str(e)}"
      )

  @staticmethod
  def export_to_csv(parent_widget, table_widget, title="Selling_Report"):
    """Экспорт данных из QTableWidget в формат CSV."""
    path, _ = QFileDialog.getSaveFileName(
        parent_widget,
        "Сохранить в CSV",
        f"{title}.csv",
        "CSV Files (*.csv)",
    )

    if not path:
      return

    try:
      with open(path, mode="w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file, delimiter=";")

        # Заголовки
        headers = []
        for col in range(table_widget.columnCount()):
          header_text = table_widget.horizontalHeaderItem(col)
          headers.append(header_text.text() if header_text else f"Col {col+1}")
        writer.writerow(headers)

        # Строки
        for row in range(table_widget.rowCount()):
          row_data = []
          for col in range(table_widget.columnCount()):
            widget = table_widget.cellWidget(row, col)
            item = table_widget.item(row, col)

            if widget and hasattr(widget, "currentText"):
              row_data.append(widget.currentText())
            elif item is not None:
              row_data.append(item.text())
            else:
              row_data.append("")

          writer.writerow(row_data)

      QMessageBox.information(
          parent_widget,
          "Успех",
          f"CSV-файл успешно сохранён:\n{path}",
      )

    except Exception as e:
      QMessageBox.critical(
          parent_widget, "Ошибка экспорта", f"Не удалось сохранить файл:\n{str(e)}"
      )